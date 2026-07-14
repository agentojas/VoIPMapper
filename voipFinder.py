import pandas as pd
import socket
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ==========================================
# CONFIGURATION
# ==========================================

INPUT_FILE = "IPDR.xlsx"
OUTPUT_FILE = "VOIP_Calls.xlsx"

DEST_IP_COL = "Destination IP"
DEST_PORT_COL = "Destination Port"

# ==========================================
# Reverse DNS Cache
# ==========================================

# For Twitter and snapchat

twitter_urls = [
    "twimg.com",
    "pbs.twimg.com",
    "video.twimg.com",
    "abs.twimg.com",
    "dualstack.twimg.twitter.map.fastly.net",
    "s.twitter.com",
    "analytics.twitter.com",
    "api.x.com",
    "ads-api.x.com",
    "chat-ws.x.com"
]

snapchat_urls = [
    "app-analytics-v2.snapchat.com",
    "aws.duplex.snapchat.com",
    "us-east1-aws.api.snapchat.com",
    "aws-proxy-gcp.api.snapchat.com",
    "us-east4-gcp.api.snapchat.com",
    "us-central1-gcp.api.snapchat.com",
    "aws.api.snapchat.com",
    "gcp.api.snapchat.com",
    "sc-gw.com",
    "gcp.api.sc-gw.com",
    "usc1-gcp-v62.api.snapchat.com"
]

zangi_urls = [
    "zvx6.zangi.com"
]


signal_urls = [
    "chat.signal.org",
"cdn3.signal.org",
"updates2.signal.org",
"cdn.signal.org",
"grpc.chat.signal.org",
"storage.signal.org"
]


wire_urls = [

    "wire.count.ly",
    "prod-nginz-https.wire.com",
    "prod-nginz-ssl.wire.com",
    "coturn-0.coturn.calling-prod-v01.wire.com",
    "coturn-1.coturn.calling-prod-v01.wire.com",
    "coturn-2.coturn.calling-prod-v01.wire.com",
    "coturn-3.coturn.calling-prod-v01.wire.com"

]

# ============================================
# Nslookup Process here
#=========================================

def resolve_domains(domains):
    ips = set()

    for domain in domains:
        try:
            _, _, iplist = socket.gethostbyname_ex(domain)
            ips.update(iplist)
        except:
            pass

    return ips


 # ----------------------------------------------------------------------------------------------------------------------



twitter_ips = resolve_domains(twitter_urls)
snapchat_ips = resolve_domains(snapchat_urls)
zangi_ips = resolve_domains(zangi_urls)
signal_ips = resolve_domains(signal_urls)
wire_ips = resolve_domains(wire_urls)

print("Twitter IPs:", twitter_ips)
print("Snapchat IPs:", snapchat_ips)
print("Zangi IPs:", zangi_ips)
print("Signal IPs:", signal_ips)
print("Wire IPs:", wire_ips)




# end declaration ---------------------------------------------------------------------------------------------------------------------



# below we check for previuos IPs last 30 hit Ips

def previous_ip_match(df, current_index, ip_set, lookback=20):  # 20 Ip back if you want increse it for better result
    start = max(0, current_index - lookback)

    for i in range(start, current_index):
        prev_ip = str(df.iloc[i]["Destination IP"]).strip()

        if prev_ip in ip_set:
            return True

    return False



# last 20 Ip checking Done -----------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------pink


# ============================================
# Nslookup Process here
#=========================================

dns_cache = {}

def reverse_dns(ip):
    ip = str(ip).strip()

    if ip in dns_cache:
        return dns_cache[ip]

    try:
        hostname = socket.gethostbyaddr(ip)[0].lower()
    except:
        hostname = ""

    dns_cache[ip] = hostname
    return hostname


# ==========================================
# Read Excel
# ==========================================

df = pd.read_excel(INPUT_FILE)

results = []

for index, row in df.iterrows():

    try:
        ip = str(row[DEST_IP_COL]).strip()
        port = int(float(row[DEST_PORT_COL]))
    except:
        continue

    # Process only VoIP related ports
    if not (port == 3478 or port == 1400 or (590 <= port <= 599)):
        continue

    hostname = reverse_dns(ip)

    # Detect Organization
    org = ""


# Checking Voip Which use 3478 Port  -----------------------------------------------------------------------------------

    if port == 3478:

        if ip.startswith("13.200."):

            if previous_ip_match(df, index, twitter_ips):
                org = "Twitter Call"

            elif previous_ip_match(df, index, snapchat_ips):
                org = "Snapchat Call"

            else:
                org = "Unknown 13.200 VOIP"


        if ip.startswith("141.101."):
            if previous_ip_match(df, index, signal_ips):
                org = "Signal Messenger Call"

            elif previous_ip_match(df, index, signal_ips):
                org = "Signal Messenger Call"

            else:
                org = "Unknown 141.101 VOIP"


        if ip.startswith("18.199"):
            if previous_ip_match(df, index, wire_ips):
                org = "Wire Messenger Call"

            else:
                org = "Unknown 18.199 VOIP"


        elif ip.startswith("162.159."):
            org = "Discord Msg / Discord Call"

       
        elif ip in zangi_ips:
            org = "Zangi Call"


        elif (
            "whatsapp" in hostname
            or "whatsapp.net" in hostname
            or "edge-stun" in hostname
            or "wasp" in hostname
    ):
            org = "WhatsApp Call"

        elif (
            "instagram" in hostname
            or "insta" in hostname
    ):
            org = "Instagram Call"

        elif port == 1400 or (590 <= port <= 599):
            org = "Telegram Call"

    new_row = row.copy()
    new_row["Reverse DNS"] = hostname
    new_row["Organization"] = org

    results.append(new_row)

# ==========================================
# Save Excel
# ==========================================

if len(results) == 0:
    print("No VoIP records found.")
    exit()

output = pd.DataFrame(results)
output.to_excel(OUTPUT_FILE, index=False)

# ==========================================
# Load Workbook
# ==========================================

wb = load_workbook(OUTPUT_FILE)
ws = wb.active

# ==========================================
# Colors Painting
# ==========================================

pink_fill = PatternFill(
    fill_type="solid",
    start_color="FFC0CB",
    end_color="FFC0CB"
)

green_fill = PatternFill(
    fill_type="solid",
    start_color="92D050",
    end_color="92D050"
)

blue_fill = PatternFill(
    fill_type="solid",
    start_color="5B9BD5",
    end_color="5B9BD5"
)

discord_fill = PatternFill(
    fill_type="solid",
    start_color="1F4E78",
    end_color="1F4E78"
)

black_fill = PatternFill(
    fill_type="solid",
    start_color="000000",
    end_color="000000"
)

yellow_fill = PatternFill(
    fill_type="solid",
    start_color="FFFF00",
    end_color="FFFF00"
)

grey_fill = PatternFill(
    fill_type="solid",
    start_color="9C9C9C",
    end_color="9C9C9C"
)

red_fill = PatternFill(
    fill_type="solid",
    start_color="FF0000",
    end_color="FF0000"
)

white_font = Font(color="FFFFFF")

# ==========================================
# Get Column Index
# ==========================================

headers = {}

for cell in ws[1]:
    headers[cell.value] = cell.column

port_col = headers[DEST_PORT_COL]
dns_col = headers["Reverse DNS"]
org_col = headers["Organization"]

# ==========================================
# Highlight Rows
# ==========================================

for row in ws.iter_rows(min_row=2):

    try:
        port = int(float(row[port_col - 1].value))
    except:
        continue

    hostname = str(row[dns_col - 1].value).lower()
    organization = str(row[org_col - 1].value)

    # ---------------------------------
    # Telegram
    # ---------------------------------

    if port == 1400 or (590 <= port <= 599):

        for cell in row:
            cell.fill = blue_fill
            cell.font = white_font

        continue

    # ---------------------------------
    #  Row Painting color Declare here
    # ---------------------------------

    if port == 3478:

        # Twitter
        if organization == "Twitter Call":

            for cell in row:
                cell.fill = black_fill
                cell.font = white_font

        # Discord
        elif organization == "Discord Msg / Discord Call":

            for cell in row:
                cell.fill = discord_fill
                cell.font = white_font

        elif organization == "Snapchat Call":

            for cell in row:
                cell.fill = yellow_fill
                

        # WhatsApp
        elif organization == "WhatsApp Call":

            for cell in row:
                cell.fill = green_fill

        # Instagram
        elif organization == "Instagram Call":

            for cell in row:
                cell.fill = pink_fill

        elif organization == "wire Messenger Call":

            for cell in row:
                cell.fill = grey_fill
                cell.font = white_font

        elif organization == "Signal Messenger Call":

            for cell in row:
                cell.fill = red_fill
                cell.font = white_font

# ==========================================
# Save Workbook
# ==========================================

wb.save(OUTPUT_FILE)

print("\n====================================")
print("VoIP Analysis Completed")
print("====================================")
print(f"Records Found : {len(results)}")
print(f"Output File   : {OUTPUT_FILE}")
